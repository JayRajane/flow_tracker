# Create your views here.
import csv
import io
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import FlowData
from .tasks import generate_flow_data
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.utils import get_column_letter

def index(request):
    return render(request, 'flow_app/index.html')

def data_view(request):
    flow_data = FlowData.objects.all().order_by('created_at')
    context = {
        'flow_data': flow_data,
    }
    return render(request, 'flow_app/data_view.html', context)

@csrf_exempt
def trigger_data_generation(request):
    daily_flow, total_flow = generate_flow_data()
    return JsonResponse({
        'status': 'success',
        'daily_flow': daily_flow,
        'total_flow': total_flow
    })

def export_excel(request):
    flow_data = FlowData.objects.all().order_by('created_at')
    
    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flow Data"
    
    # Add headers
    headers = ['Date', 'Time', 'Day', 'Daily Flow', 'Total Flow']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
    
    # Add data
    row_num = 2
    for data in flow_data:
        ws[f'A{row_num}'] = data.date.strftime('%d-%m-%Y')
        ws[f'B{row_num}'] = data.time.strftime('%H:%M')
        ws[f'C{row_num}'] = data.day
        ws[f'D{row_num}'] = data.daily_flow
        ws[f'E{row_num}'] = data.total_flow
        row_num += 1
    
    # Set column widths
    for col_num in range(1, 6):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=flow_data.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response

def export_pdf(request):
    flow_data = FlowData.objects.all().order_by('created_at')
    
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object, using the buffer as its "file"
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Create a list of data for the table
    data = [['Date', 'Time', 'Day', 'Daily Flow', 'Total Flow']]
    
    for item in flow_data:
        data.append([
            item.date.strftime('%d-%m-%Y'),
            item.time.strftime('%H:%M'),
            item.day,
            str(item.daily_flow),
            str(item.total_flow)
        ])
    
    # Create the table
    table = Table(data)
    
    # Style the table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    
    table.setStyle(style)
    
    # Add the table to the PDF document
    elements = []
    elements.append(table)
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=flow_data.pdf'
    response.write(pdf)
    
    return response